import re
import tempfile
from datetime import date, datetime
from pathlib import Path

from django.conf import settings
from django.contrib.auth.models import AbstractUser
from django.db import models
from django.templatetags import tz
from django.utils import timezone
from django.utils.safestring import mark_safe
from django.utils.translation import gettext as _

from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell
from openpyxl.styles import numbers
from openpyxl.workbook.child import INVALID_TITLE_REGEX
from openpyxl.worksheet.worksheet import Worksheet

from keyta.apps.executions.errors import ValidationError
from keyta.models.base_model import AbstractBaseModel
from keyta.widgets import html_to_string, Icon, link

from ..types import (
    ParamData,
    ParamMetadata,
    StepData,
    StepMetadata,
    StepParameterValues,
    TestStepsData
)


def as_date(value: str) -> datetime|None:
    try:
        return datetime.strptime(value, '%d.%m.%Y')
    except ValueError:
        return None


def autosize_worksheet_columns(worksheet: Worksheet):
    for column_cells in worksheet.columns:
        max_width = (max(len(str(cell.value)) or 10 for cell in column_cells) + 2)
        worksheet.column_dimensions[column_cells[0].column_letter].width = max_width


def format_cells(worksheet: Worksheet):
    for column in worksheet.columns:
        for cell in column:
            if cell.value:
                if date := as_date(cell.value):
                    cell.value = date
                    cell.data_type = 'd'
                    cell.number_format = 'dd.mm.yyyy'
                elif cell.value.startswith('0') and len(cell.value) > 1:
                    pass
                elif is_int(cell.value):
                    cell.data_type = 'n'
                    cell.number_format = '0'
                elif is_int(normalize(cell.value)):
                    cell.data_type = 'n'
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                    value = cell.value.replace('.', '')

                    if re.search(r',\d\d?$', cell.value):
                        cell.value = float(value.replace(',', '.'))
                    else:
                        cell.value = int(value)
                else:
                    pass


def format_value(value: str):
    if re.search(r'\.\d\d?$', value):
        return value.replace('.', ',')

    return value


def get_cell_content(cell: Cell):
    if cell.value is None:
        return ''

    if isinstance(cell.value, date):
        return cell.value.strftime("%d.%m.%Y")
    else:
        return str(cell.value)


def get_excel_file_path(testdata_name: str):
    tmp_dir = Path(tempfile.gettempdir()) / 'KeyTA'
    tmp_dir.mkdir(exist_ok=True)

    return tmp_dir / f'{testdata_name}.xlsx'


def is_int(value: str) -> bool:
    try:
        int(value)
        return True
    except ValueError:
        return False


def merge(original: TestStepsData, patch: TestStepsData) -> TestStepsData:
    patch_steps_pk_to_index = {}
    patch_to_keyword_pks = set()

    for index, step_metadata in enumerate(patch['metadata']):
        patch_steps_pk_to_index.update({step_metadata['pk']: index})
        patch_to_keyword_pks.add(step_metadata['to_keyword_pk'])

    for index, step_metadata in enumerate(original['metadata']):
        pk = step_metadata['pk']
        to_keyword_pk = step_metadata['to_keyword_pk']

        if pk not in patch_steps_pk_to_index or to_keyword_pk not in patch_to_keyword_pks:
            original['steps'].pop(index)
            original['metadata'].pop(index)

    orig_steps_pk_to_index = {}
    orig_to_keyword_pks = set()

    for index, step_metadata in enumerate(original['metadata']):
        orig_steps_pk_to_index.update({step_metadata['pk']: index})
        orig_to_keyword_pks.add(step_metadata['to_keyword_pk'])

    for patch_index, patch_step_metadata in enumerate(patch['metadata']):
        pk = patch_step_metadata['pk']
        to_keyword_pk = patch_step_metadata['to_keyword_pk']
        patch_step = patch['steps'][patch_index]

        if pk not in orig_steps_pk_to_index or to_keyword_pk not in orig_to_keyword_pks:
            original['steps'].append(patch_step)
            original['metadata'].append(patch_step_metadata)
        else:
            orig_index = orig_steps_pk_to_index[pk]
            orig_type = original['metadata'][orig_index]['type']
            original['steps'][orig_index]['index'] = patch_step['index']
            original['steps'][orig_index]['name'] = patch_step['name']
            original['metadata'][orig_index]['index'] = patch_step['index']
            patch_type = patch_step_metadata['type']

            if orig_type != patch_type:
                original['steps'][orig_index]['params'] = patch['steps'][patch_index]['params']
                original['metadata'][orig_index]['params'] = patch['metadata'][patch_index]['params']
                original['metadata'][orig_index]['type'] = patch_type
            else:
                if orig_type == 'DICT':
                    merge_dicts(original, patch, orig_index, patch_index)
                if orig_type == 'LIST':
                    merge_lists(original, patch, orig_index, patch_index)

    original['steps'] = sorted(original['steps'], key=lambda step: step['index'])
    original['metadata'] = sorted(original['metadata'], key=lambda step: step['index'])

    return original


def merge_dicts(
    original: TestStepsData,
    patch: TestStepsData,
    orig_index: int,
    patch_index: int
):
    patch_params_pk_to_index = {
        param_metadata['pk']: index
        for index, param_metadata in enumerate(patch['metadata'][patch_index]['params'])
    }

    for index, param_metadata in enumerate(original['metadata'][orig_index]['params']):
        pk = param_metadata['pk']

        if pk not in patch_params_pk_to_index:
            original['steps'][orig_index]['params'].pop(index)
            original['metadata'][orig_index]['params'].pop(index)

    orig_params_pk_to_index = {
        param_metadata['pk']: index
        for index, param_metadata in enumerate(original['metadata'][orig_index]['params'])
    }

    for param_metadata in patch['metadata'][patch_index]['params']:
        pk = param_metadata['pk']
        patch_param_index = patch_params_pk_to_index[pk]
        patch_param = patch['steps'][patch_index]['params'][patch_param_index]
        patch_param_metadata = patch['metadata'][patch_index]['params'][patch_param_index]

        if pk not in orig_params_pk_to_index:
            original['steps'][orig_index]['params'].append(patch_param)
            original['metadata'][orig_index]['params'].append(patch_param_metadata)
        else:
            orig_param_index = orig_params_pk_to_index[pk]
            original['steps'][orig_index]['params'][orig_param_index]['index'] = patch_param['index']
            original['steps'][orig_index]['params'][orig_param_index]['name'] = patch_param['name']
            original['metadata'][orig_index]['params'][orig_param_index]['index'] = patch_param['index']

    original['steps'][orig_index]['params'] = sorted(original['steps'][orig_index]['params'], key=lambda param: param['index'])
    original['metadata'][orig_index]['params'] = sorted(original['metadata'][orig_index]['params'], key=lambda param: param['index'])


def merge_lists(
    original: TestStepsData,
    patch: TestStepsData,
    orig_index: int,
    patch_index: int
):
    orig_column_metadata = original['metadata'][orig_index]['params']
    patch_column_metadata = patch['metadata'][patch_index]['params']

    if len(orig_column_metadata) == len(patch_column_metadata):
        orig_column_order = {
            index: column['pk']
            for index, column in enumerate(orig_column_metadata)
        }
        patch_column_order = {
            column['pk']: index
            for index, column in enumerate(patch_column_metadata)
        }
        update_column_order = {
            index: patch_column_order[pk]
            for index, pk in orig_column_order.items()
        }
        _orig_column_titles, *orig_rows = original['steps'][orig_index]['params']
        patch_column_titles, *_patch_rows = patch['steps'][patch_index]['params']
        ordered_rows = [
            [
                col
                for index, col in
                sorted(enumerate(row), key=lambda index_col: update_column_order[index_col[0]])
            ]
            for row in orig_rows
        ]
        ordered_table = [
            patch_column_titles,
            *ordered_rows
        ]
        original['steps'][orig_index]['params'] = ordered_table
        original['metadata'][orig_index]['params'] = patch['metadata'][patch_index]['params']
    else:
        patch_columns_pk_to_index = {
            column_metadata['pk']: index
            for index, column_metadata in enumerate(patch['metadata'][patch_index]['params'])
        }

        for column_index, original_column_metadata in enumerate(original['metadata'][orig_index]['params']):
            pk = original_column_metadata['pk']

            if pk not in patch_columns_pk_to_index:
                column_titles, *rows = original['steps'][orig_index]['params']
                column_titles.pop(column_index)
                row: list
                for row in rows:
                    row.pop(column_index)
                original['steps'][orig_index]['params'] = [column_titles, *rows]
                original['metadata'][orig_index]['params'].pop(column_index)

        orig_columns_pk_to_index = {
            column_metadata['pk']: index
            for index, column_metadata in enumerate(original['metadata'][orig_index]['params'])
        }

        for patch_column_metadata in patch['metadata'][patch_index]['params']:
            pk = patch_column_metadata['pk']
            column_index = patch_columns_pk_to_index[pk]

            if pk not in orig_columns_pk_to_index:
                orig_column_titles, *orig_rows = original['steps'][orig_index]['params']
                patch_column_titles, *patch_rows = patch['steps'][patch_index]['params']
                patch_column = [
                    row[column_index]
                    for row in patch_rows
                ]
                orig_column_titles.insert(column_index, patch_column_titles[column_index])

                for r, row in enumerate(orig_rows):
                    row.insert(column_index, patch_column[r])

                original['steps'][orig_index]['params'] = [orig_column_titles, *orig_rows]
                original['metadata'][orig_index]['params'].append(patch_column_metadata)
                original['metadata'][orig_index]['params'] = sorted(original['metadata'][orig_index]['params'], key=lambda metadata: metadata['index'])


def normalize(value: str) -> str:
    return (
        value
        .replace('.', '')
        .replace(',', '')
    )


class TestDataError(Exception):
    pass


def read_from_excel(filename: str, test_steps_metadata: list[StepMetadata]) -> list[StepData]:
    worksheets = [ws for ws in load_workbook(filename)]
    result = []

    if len(worksheets) != len(test_steps_metadata):
        raise TestDataError()

    for ws, step_metadata in zip(worksheets, test_steps_metadata):
        step_index, step_name = ws.title.split('|')
        step_data: StepData = {
            'index': int(step_index),
            'name': step_name,
            'params': []
        }
        table = worksheet_as_list_of_lists(ws)

        if step_metadata['type'] == 'DICT':
            if len(table) != len(step_metadata['params']):
                raise TestDataError()

            step_data['params'] = table_to_param_data(table)

        if step_metadata['type'] == 'LIST':
            first_row, *_ = table

            if len(first_row) != len(step_metadata['params']):
                raise TestDataError()

            step_data['params'] = table_to_table_data(table)

        result.append(step_data)

    return result


def save_to_excel(test_steps_data: TestStepsData, filename: str):
    wb = Workbook()
    if active_ws := wb.active:
        wb.remove(active_ws)

    for index, step in enumerate(test_steps_data['steps']):
        step_index = step['index']
        step_name = step['name']
        step_params = step['params']

        # Excel worksheet names must be unique
        title = f'{step_index}|{step_name}'
        # Excel limits worksheet names to 31 chars
        title = title[:31]
        # Excel worksheet names cannot include :, /, ?, *, [, ]
        title = INVALID_TITLE_REGEX.sub('_', title)
        ws = wb.create_sheet(title)

        if test_steps_data['metadata'][index]['type'] == 'DICT':
            table = [
                [param['name'], param['value']]
                for param in step_params
            ]
            for row in table:
                ws.append(row)

        if test_steps_data['metadata'][index]['type'] == 'LIST':
            column_headings, *table = step_params
            ws.append([col['name'] for col in column_headings])
            for row in table:
                ws.append(row)

        autosize_worksheet_columns(ws)
        format_cells(ws)

    wb.save(filename)


def serialize_row(row: list):
    return [get_cell_content(cell) for cell in row]


def table_to_param_data(table: list) -> list[ParamData]:
    return [
        {
            'index': int(index),
            'name': name,
            'value': format_value(value)
        }
        for index, (name, value) in enumerate(table)
    ]


def table_to_table_data(table: list) -> list[list]:
    column_titles, *rows = table

    return [
        [
            {
                'index': int(index),
                'name': column
            }
            for index, column in enumerate(column_titles)
        ],
        *[
            [format_value(col) for col in row]
            for row in rows
        ]
    ]


def worksheet_as_list_of_lists(worksheet: Worksheet):
    return [
        serialize_row(row)
        for row in worksheet.iter_rows()
    ]


class TestData(AbstractBaseModel):
    testcase = models.ForeignKey(
        'testcases.TestCase',
        on_delete=models.CASCADE,
        related_name='testdata'
    )
    data = models.JSONField(
        default=dict
    )
    last_update = models.DateTimeField(
        null=True,
        verbose_name=_('Letzte Aktualisierung')
    )
    name = models.CharField(
        max_length=255
    )

    def __str__(self):
        return self.name

    def export_to_excel(self, user: AbstractUser):
        self.update(self.testcase.get_test_steps_data(user))
        excel_file_path = get_excel_file_path(self.name)
        save_to_excel(self.data, str(excel_file_path))
        return excel_file_path

    def get_download_icon(self):
        if not self.pk:
            return '-'

        return link(
            self.get_tab_url(),
            str(Icon(
                settings.FA_ICONS.export_testdata,
                {'font-size': '1.5em', 'margin-top': '5px'}
            )),
            attrs={
                'export-url': self.get_admin_url() + '?export',
                'id': f'export-testdata-{self.pk}'
            }
        )

    def get_file_input(self):
        if not self.pk:
            return ''

        return mark_safe(html_to_string(
            'input',
            {
                'accept': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'id': f'excel-file-{self.pk}',
                'import-url': self.get_admin_url(),
                'style': 'display: none',
                'type': 'file'
            },
            '',
            self_closing=True
        ))

    def get_last_update_field(self):
        if self.last_update:
            last_update = tz.localtime(self.last_update).strftime('%d.%m.%Y %H:%M:%S')
        else:
            last_update = '-'

        return mark_safe(html_to_string(
            'span',
            {
                'hx-get': self.get_admin_url() + '?last_update',
                'hx-swap': 'outerHTML',
                'hx-target': 'this',
                'hx-trigger': f'update-last-update-{self.pk} from:body'
            },
            last_update
        ))

    def get_metadata(self, user: AbstractUser):
        return self.testcase.get_test_steps_data(user)['metadata']

    def get_parameter_values(self) -> dict[int, StepParameterValues]:
        data: TestStepsData = self.data
        result = {}

        step: StepData
        metadata: StepMetadata
        for step, metadata in zip(data['steps'], data['metadata']):
            params: dict[int, str] = {}
            table = None

            if metadata['type'] == 'DICT':
                param: ParamData
                param_metadata: ParamMetadata
                for param, param_metadata in zip(step['params'], metadata['params']):
                    params[param_metadata['pk']] = param['value']

            if metadata['type'] == 'LIST':
                column_headings, *rows = step['params']
                column_names = [col['name'] for col in column_headings]
                table_name = f"{step['index']}|{step['name']}"
                table = (table_name, [column_names, *rows])

            result[metadata['pk']] = {
                'params': params,
                'table': table
            }

        return result

    def get_upload_icon(self):
        if not self.pk:
            return '-'

        if not self.data:
            cursor = 'default'
            title = '-'
        else:
            cursor = 'pointer'
            title = str(Icon(
                settings.FA_ICONS.import_testdata,
                {'font-size': '1.5em', 'margin-top': '5px'}
            ))

        return link(
            self.get_tab_url(),
            title,
            attrs={
                'hx-on:click': f"document.getElementById('excel-file-{self.pk}').click()",
                'hx-get': self.get_admin_url() + '?upload_icon',
                'hx-swap': 'outerHTML',
                'hx-target': 'this',
                'hx-trigger': f'show-upload-icon-{self.pk} from:body'
            },
            styles={'cursor': cursor}
        )

    def import_from_excel(self, excel_file_path: Path, metadata: list[StepMetadata]):
        self.data['steps'] = read_from_excel(str(excel_file_path), metadata)
        self.data['metadata'] = metadata
        self.last_update = timezone.now()
        self.save()

    def update(self, test_steps_data: TestStepsData):
        if self.data:
            self.data = merge(self.data, test_steps_data)
        else:
            self.data = test_steps_data

        self.save()

    def validate_metadata(self, metadata: list[StepMetadata]):
        saved_metadata = self.data['metadata']

        if len(metadata) != len(saved_metadata):
            return ValidationError.INVALID_TESTDATA

        for step_metadata, saved_step_metadata in zip(metadata, saved_metadata):
            if len(step_metadata['params']) != len(saved_step_metadata['params']):
                return ValidationError.INVALID_TESTDATA

            if step_metadata['type'] != saved_step_metadata['type']:
                return ValidationError.INVALID_TESTDATA

    class Meta:
        ordering = ['name']
        verbose_name = _('Testdaten')
        verbose_name_plural = _('Testdaten')
